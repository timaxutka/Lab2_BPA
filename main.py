import openpyxl
from aiogram import Bot, types
from aiogram.dispatcher import Dispatcher
from aiogram.utils import executor
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
TOKEN = '5973296887:AAEvRGAl9R8vqMWVt9xTBcnQtj9MESLtDsc'

bot = Bot(token=TOKEN)
dp = Dispatcher(bot)

kb = ReplyKeyboardMarkup(resize_keyboard=True)
b1 = KeyboardButton(text="/choose")
b2 = KeyboardButton(text="/students")
kb.add(b1, b2)

@dp.message_handler(commands=['start'])
async def process_start_command(message: types.Message):
    await bot.send_message(chat_id=message.from_user.id,  text="Привет!\nЯ чат бот для студентов по дисциплине!\n"
                        "Выбери студента, а я скажу тебе его посещаемость и баллы!\n"
                         "Вот, что я могу:\n\n1. Узнать список студентов: /students\n\n"
                        "2. Узнать количество пропусков и баллов отдельного студента: /choose [Имя_Студента]", reply_markup=kb)

@dp.message_handler(commands=['students'])
async def get_students_list(message: types.Message):
    book = openpyxl.load_workbook('посещаемость.xlsx') # Открываем файл Excel и получаем нужный лист
    sheet = book.active

    search_list = {}
    current_number = 4
    current_cell = sheet['B' + str(current_number)]

    rows = sheet.max_row
    for i in range(1, rows - 2): # Получаем список студентов из колонны "B"
        search_list[i] = current_cell.value
        current_cell = sheet['B']
        current_number += 1
        current_cell = sheet['B' + str(current_number)]
    list_data = [] # перевод из словаря в список
    for k, v in search_list.items():
        list_data.append(str(v))
    await message.answer('Список студентов:\n\n' + '\n'.join(list_data)) # Отправляем список студентов в ответном сообщении
    book.close()

@dp.message_handler(commands=['choose'])
async def get_student_info(message: types.Message):
    book = openpyxl.load_workbook("посещаемость.xlsx")
    sheet = book.active
    absences, student_name = message.text.split()


    for row in sheet.iter_rows(min_row=4, max_row=21, min_col=2, max_col=2):
        if row[0].value == student_name:
            absences = sheet.cell(row=row[0].row, column=34).value # получение количества пропусков и баллов
            points = sheet.cell(row=row[0].row, column=35).value
            await message.answer(f"Студент {student_name}\nКоличество пропусков: {absences}\nКоличество баллов {points}")
            book.close()
            return
    await message.answer(f"Студент {student_name} не найден")
    book.close()

if __name__ == '__main__':
    executor.start_polling(dp)